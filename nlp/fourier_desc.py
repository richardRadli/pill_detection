import colorama
import cv2
import gc
import json
import matplotlib.pyplot as plt
import numpy as np
import shutil
import pyefd
import os
import openpyxl

from concurrent.futures import ProcessPoolExecutor
from glob import glob
from multiprocessing import Manager
from tqdm import tqdm
from scipy.spatial.distance import directed_hausdorff, pdist, squareform, cosine
from sklearn.decomposition import PCA

from config.config_selector import Fourier_configs, dataset_images_path_selector
from utils.utils import create_timestamp, find_latest_file_in_directory, measure_execution_time


class NumpyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        return super(NumpyEncoder, self).default(obj)


class FourierDescriptor:
    def __init__(self, copy_images, load, order):
        self.timestamp = create_timestamp()
        colorama.init()

        self.order = order
        self.copy_images = copy_images
        self.load = load

        self.images_dir = dataset_images_path_selector().get("nih").get("ref")
        self.file_path = Fourier_configs().get("Fourier_collected_images_by_shape_nih")
        self.plot_efd_dir = Fourier_configs().get("Fourier_plot_shape")
        self.plot_euc_dir = Fourier_configs().get("Fourier_euclidean_distance")
        self.json_dir = Fourier_configs().get("Fourier_saved_mean_vectors")
        self.excel_path = os.path.join(dataset_images_path_selector().get("nih").get("xlsx"), "ref.xlsx")

    def get_excel_values(self) -> None:
        """
        Reads an Excel file, processes the data, and copies corresponding files based on specified conditions.

        :return: None
        """

        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook['Sheet1']

        shape_dict = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            ndc11 = row[0]
            shape = row[5]

            if shape in shape_dict:
                shape_dict[shape].append(ndc11)
            else:
                shape_dict[shape] = [ndc11]

        workbook.close()

        for shape, ndc11_list in tqdm(shape_dict.items(), desc=colorama.Fore.GREEN + "Processing Shapes"):
            print(f"Shape: {shape}")
            for ndc11 in ndc11_list:
                src_dir = os.path.join(self.images_dir, str(ndc11))

                try:
                    files = os.listdir(src_dir)[0]
                    src_file = os.path.join(src_dir, files)
                    dst_dir = os.path.join(self.file_path, shape)
                    os.makedirs(dst_dir, exist_ok=True)
                    shutil.copy(src_file, dst_dir)
                except FileNotFoundError:
                    print(f"File not found: {src_file}")

    def plot_efd(self, filename: str, coeffs: np.ndarray, locus: tuple = (0.0, 0.0), image: np.ndarray = None,
                 contour: np.ndarray = None, n: int = 300) -> None:
        """
         Plots the elliptical Fourier descriptors.

         :param filename: Name of the file to save the plot.
         :param coeffs: Coefficients for elliptical Fourier descriptors.
         :param locus: Tuple representing the locus (center) of the plot.
         :param image: Background image (optional).
         :param contour: Contour data (optional).
         :param n: Number of points to generate the plot.
         :return: None
         """

        num_of_coeffs = coeffs.shape[0]
        n_half = int(np.ceil(num_of_coeffs / 2))
        n_rows = 2

        t = np.linspace(0, 1.0, n)
        xt = np.ones((n,)) * locus[0]
        yt = np.ones((n,)) * locus[1]

        for n in range(coeffs.shape[0]):
            xt += (coeffs[n, 0] * np.cos(2 * (n + 1) * np.pi * t)) + (
                    coeffs[n, 1] * np.sin(2 * (n + 1) * np.pi * t)
            )
            yt += (coeffs[n, 2] * np.cos(2 * (n + 1) * np.pi * t)) + (
                    coeffs[n, 3] * np.sin(2 * (n + 1) * np.pi * t)
            )
            ax = plt.subplot2grid((n_rows, n_half), (n // n_half, n % n_half))
            ax.set_title(str(n + 1))

            if image is not None:
                # A background image of shape [rows, cols] gets transposed by imshow so that the first dimension is
                # vertical and the second dimension is horizontal. This implies swapping the x and y axes when plotting
                # a curve.
                if contour is not None:
                    ax.plot(contour[:, 1], contour[:, 0], "c--", linewidth=2)
                ax.plot(yt, xt, "r", linewidth=2)
                ax.imshow(image, cmap="gray")
            else:
                # Without a background image, no transpose is implied. This case is useful when (x,y) point clouds
                # without relation to an image are to be handled.
                if contour is not None:
                    ax.plot(contour[:, 0], contour[:, 1], "c--", linewidth=2)
                ax.plot(xt, yt, "r", linewidth=2)
                ax.axis("equal")

        # plt.show()
        plt.tight_layout()
        plt.savefig(os.path.join(self.plot_efd_dir, f"{filename}"))
        plt.close()
        gc.collect()

    @staticmethod
    def preprocess_image(image: np.ndarray) -> np.ndarray:
        """
        Preprocesses an input image by applying edge detection, thresholding, dilation, and connected components
        analysis.

        :param image: Input image.
        :return: Preprocessed image.
        """
        sobel_x = cv2.Sobel(image, cv2.CV_64F, 1, 0, ksize=3)
        sobel_y = cv2.Sobel(image, cv2.CV_64F, 0, 1, ksize=3)
        edge_map = cv2.add(np.abs(sobel_x), np.abs(sobel_y))

        _, thresholded_edge_map = cv2.threshold(edge_map, 10, 255, cv2.THRESH_BINARY)
        dilated_edge_map = cv2.dilate(thresholded_edge_map, kernel=np.ones((5, 5), np.uint8), iterations=1)
        dilated_edge_map = dilated_edge_map.astype(np.uint8)
        _, labels, stats, centroids = cv2.connectedComponentsWithStats(dilated_edge_map)

        sorted_indices = np.argsort(stats[:, 4])[::-1]
        second_largest_index = sorted_indices[1]

        pill_mask = np.zeros_like(dilated_edge_map)
        pill_mask[labels == second_largest_index] = 255

        return cv2.bitwise_and(image, image, mask=pill_mask)

    @staticmethod
    def get_largest_contour(segmented_image: np.ndarray):
        contours, _ = cv2.findContours(segmented_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        largest_contour = max(contours, key=cv2.contourArea)
        return np.squeeze(largest_contour)

    def elliptic_fourier_wo_norm(self, image: np.ndarray, segmented_image: np.ndarray, filename: str = None):
        """
        Compute elliptic Fourier descriptors for the largest contour in the segmented image.

        :param image: Original image.
        :param segmented_image: Segmented image.
        :param filename: Name of the file for plot (optional).
        :return: Optional
        """

        largest_contour = self.get_largest_contour(segmented_image)
        number_of_points = largest_contour.shape[0]
        locus = pyefd.calculate_dc_coefficients(largest_contour)
        coefficients = pyefd.elliptic_fourier_descriptors(largest_contour, order=self.order, normalize=False)

        image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        rotated_image = cv2.transpose(image)
        self.plot_efd(filename, coefficients, locus, rotated_image, largest_contour)

        reconstruction = pyefd.reconstruct_contour(coefficients, locus, number_of_points)
        hausdorff_distance, _, _ = directed_hausdorff(reconstruction, largest_contour)
        print(hausdorff_distance)

    def elliptic_fourier_w_norm(self, segmented_image: np.ndarray):
        largest_contour = self.get_largest_contour(segmented_image)
        coefficients = pyefd.elliptic_fourier_descriptors(largest_contour, order=self.order, normalize=True)
        np.testing.assert_almost_equal(coefficients[0, 0], 1.0, decimal=14)
        np.testing.assert_almost_equal(coefficients[0, 1], 0.0, decimal=14)
        np.testing.assert_almost_equal(coefficients[0, 2], 0.0, decimal=14)
        return coefficients.flatten()[3:]

    def plot_euclidean_distances(self, class_averages: dict) -> None:
        """
        Plot pairwise Euclidean distances between class averages.

        :param class_averages: Dictionary mapping class labels to class averages.
        :return: None
        """

        class_labels = list(class_averages.keys())
        class_vectors = np.array(list(class_averages.values()))

        normalized_values = ((class_vectors - np.min(class_vectors, axis=0)) /
                             (np.max(class_vectors, axis=0) - np.min(class_vectors, axis=0)))

        distances = squareform(pdist(normalized_values, 'euclidean'))

        plt.figure(figsize=(8, 8))
        plt.imshow(distances, cmap='viridis', vmin=0, vmax=np.max(distances))
        plt.colorbar(label='Euclidean Distance')

        plt.xticks(np.arange(len(class_labels)), class_labels, rotation=45, ha='right')
        plt.yticks(np.arange(len(class_labels)), class_labels)
        plt.xlabel('Class Label')
        plt.ylabel('Class Label')
        plt.title('Pairwise Euclidean Distances between Class Averages')

        plt.tight_layout()
        filename = os.path.join(self.plot_euc_dir, f"euclidean_distances_{self.order}_{self.timestamp}.png")
        plt.savefig(filename, dpi=300)
        plt.close()

    @staticmethod
    def calculate_distance(vector1: list, vector2: list) -> float:
        """
        Calculate the Cosine distance between two vectors.

        :param vector1: First vector.
        :param vector2: Second vector.
        :return: Cosine distance between the two vectors.
        """

        return cosine(vector1, vector2)

    def compare_distances(self, class_averages, image_path):
        """
        Compare the Fourier descriptor of a new image with class averages and print the closest match.

        :param class_averages: Dictionary mapping class labels to class averages.
        :param image_path:
        :return: None
        """
        hit = 0

        if image_path.endswith('.JPG'):
            new_image_original = cv2.imread(image_path, 1)
            new_image = cv2.cvtColor(new_image_original, cv2.COLOR_BGR2GRAY)
            new_segmented_image = self.preprocess_image(new_image)
            new_fourier_descriptor = self.elliptic_fourier_w_norm(segmented_image=new_segmented_image)

            closest_class = None
            min_distance = float('inf')

            for class_label, class_vector in class_averages.items():
                distance = self.calculate_distance(new_fourier_descriptor, class_vector)
                # print(f"Distance to {class_label}: {distance}")

                if distance < min_distance:
                    min_distance = distance
                    closest_class = class_label

            workbook = openpyxl.load_workbook(self.excel_path)
            sheet = workbook['Sheet1']
            original_shape = []

            image_path = image_path.replace("\\", "/")
            query_ndc11 = image_path.split("/")[-2]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                ndc11 = row[0]
                shape = row[5]
                if query_ndc11 == ndc11:
                    original_shape.append(shape)

            if original_shape[0] == closest_class:
                print(colorama.Fore.LIGHTGREEN_EX +
                      f"The predicted shape {closest_class} matches the original shape {original_shape[0]}")
                hit += 1
            else:
                print(colorama.Fore.LIGHTRED_EX +
                      f"The predicted shape {closest_class} does not match the original shape {original_shape[0]}")

            return hit, original_shape[0], closest_class

    @staticmethod
    def plot_vectors(class_averages):
        classes = list(class_averages.keys())
        vectors = list(class_averages.values())

        pca = PCA(n_components=3)
        reduced_vectors = pca.fit_transform(vectors)

        plt.figure(figsize=(10, 8))
        ax = plt.axes(projection='3d')
        ax.scatter(reduced_vectors[:, 0], reduced_vectors[:, 1], reduced_vectors[:, 2],
                   c=range(len(classes)),
                   s=80,
                   marker='o',
                   cmap="viridis",
                   alpha=0.7)
        ax.set_zlim3d(-1, 1)

        for i, label in enumerate(classes):
            ax.text(reduced_vectors[i, 0], reduced_vectors[i, 1], reduced_vectors[i, 2] + 0.02,
                    label, ha='center', va='center')

        plt.title('Point Cloud After Dimensionality Reduction')
        plt.xlabel('Dimension 1')
        plt.ylabel('Dimension 2')
        plt.show()

    def save_json(self, class_averages: dict) -> None:
        """
        Save class averages to a JSON file.

        :param class_averages: Dictionary mapping class labels to class averages.
        :return: None
        """
        file_name = os.path.join(self.json_dir, f"{self.timestamp}_average-vectors_order_{self.order}.json")
        with open(file_name, "w") as json_file:
            json.dump(class_averages, json_file, cls=NumpyEncoder)

    def load_json(self) -> dict:
        """
        Load class averages from a JSON file.

        :return: Dictionary mapping class labels to class averages.
        """

        latest_file = find_latest_file_in_directory(path=self.json_dir, extension="json")
        with open(latest_file, "r") as json_file:
            class_averages = json.load(json_file)
        return class_averages

    def execute_comparison(self, class_averages: dict, image_path: str, cnt: int, label_counts: dict):
        """

        :param class_averages:
        :param image_path:
        :param cnt:
        :param label_counts:
        :return:
        """

        hit_cnt, original, predicted = self.compare_distances(class_averages, image_path)
        cnt += hit_cnt
        if predicted == original:
            label_counts[original] += 1
        return cnt, None

    @measure_execution_time
    def main(self) -> None:
        """

        :return:
        """

        if self.copy_images:
            self.get_excel_values()

        if not self.load:
            class_averages = {}

            for root, dirs, files in os.walk(self.file_path):
                for class_label in tqdm(dirs, desc=colorama.Fore.BLUE + "Processing classes"):
                    class_coefficients = []

                    class_path = str(os.path.join(root, class_label))
                    for file in tqdm(os.listdir(class_path), desc=colorama.Fore.YELLOW + "Processing vectors"):
                        image_path = os.path.join(class_path, file)
                        image_original = cv2.imread(image_path, 1)
                        try:
                            image = cv2.cvtColor(image_original, cv2.COLOR_BGR2GRAY)
                            segmented_image = self.preprocess_image(image)
                            norm_fourier_coeff = self.elliptic_fourier_w_norm(segmented_image=segmented_image)
                            class_coefficients.append(norm_fourier_coeff)
                        except Exception as e:
                            print(f"{e}, Could not open{image_path}")

                    if class_coefficients:
                        class_average = np.mean(class_coefficients, axis=0)
                        class_averages[class_label] = class_average

            self.plot_euclidean_distances(class_averages)
            self.save_json(class_averages)
        else:
            try:
                class_averages = self.load_json()
                classes = os.listdir(self.images_dir)
                cnt = 0
                class_labels = list(class_averages.keys())
                label_counts = Manager().dict({key: 0 for key in class_labels})

                # self.plot_vectors(class_averages)

                with ProcessPoolExecutor(max_workers=8) as executor:
                    features = []
                    for clas in classes:
                        image_paths = sorted(glob(self.images_dir + f"/{clas}/" + "*"))
                        for idx, image_path in enumerate(image_paths):
                            if idx == 1:
                                future = executor.submit(
                                    self.execute_comparison, class_averages, image_path, cnt, label_counts
                                )
                                features.append(future)

                    for future in features:
                        result_cnt, result_label_counts = future.result()
                        cnt += result_cnt

                    print(colorama.Fore.WHITE + "Hit ratio: ", cnt / len(classes))

                    for key, value in label_counts.items():
                        print(f"{key}: {value}")

            except FileNotFoundError as fne:
                print(f"{fne}")


if __name__ == "__main__":
    try:
        fd = FourierDescriptor(copy_images=False, load=False, order=15)
        fd.main()
    except KeyboardInterrupt:
        print("Ctrl+C pressed")
