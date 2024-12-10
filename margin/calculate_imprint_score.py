import json
import logging
import numpy as np
import os
import openpyxl

from typing import Dict, List

from config.json_config import json_config_selector
from config.dataset_paths_selector import dataset_images_path_selector
from utils.utils import create_timestamp, NumpyEncoder, load_config_json


def encoding(words: List[str], feature_type: str) -> Dict[str, List[float]]:
    """
    Encode a list of words using one-hot encoding.

    Args:
        words (List[str]): List of words to encode.
        feature_type (str): Type of feature.

    Returns:
        Dict[str, List[float]]: A dictionary where each word is paired with its one-hot encoded vector as a list of
        floats.
    """

    if feature_type == 'score_vectors':
        encoded_features = np.array([[0, 1, 1], [1, 0, 0.5], [1, 0.5, 0]])
    elif feature_type == 'imprint_vectors':
        encoded_features = np.array([[0, 0.5, 1], [0.5, 0, 1], [1, 1, 0]])
    else:
        raise ValueError(f'Wrong feature type {feature_type}')

    return {word: encoded_features[idx].tolist() for idx, word in enumerate(words)}


def create_feature_vectors(sheet, words: list, feature_type: str, dataset_name) -> dict:
    """
    Create feature vectors based on input words and feature type.

    Args:
        sheet (Workbook): Excel sheet object.
        words (list): List of words for encoding.
        feature_type (str): Type of feature vectors to use.
        dataset_name (str):
    Returns:
        dict: Dictionary containing feature vectors.
    """

    encoded_dict = encoding(words, feature_type)
    feature_dict = {}

    if dataset_name == "cure_one_sided":
        for row in sheet.iter_rows(min_row=2, values_only=True):
            pill_id = row[0]

            selected_column = row[2] if feature_type == "imprint_vectors" else row[5]
            if selected_column is None:
                selected_column = "None"

            if pill_id not in feature_dict:
                feature_dict[pill_id] = []

            encoded_feature = encoded_dict.get(selected_column)
            feature_dict[pill_id].append(encoded_feature)

    elif dataset_name == "cure_two_sided" or dataset_name == "ogyeiv2":
        for row in sheet.iter_rows(min_row=2, values_only=True):
            pill_id = row[0]
            selected_column_1 = row[3] if feature_type == "imprint_vectors" else row[7]
            selected_column_2 = row[4] if feature_type == "imprint_vectors" else row[8]

            if selected_column_1 is None:
                selected_column_1 = "None"

            if selected_column_2 is None:
                selected_column_2 = "None"

            if pill_id not in feature_dict:
                feature_dict[pill_id] = []

            encoded_feature_1 = encoded_dict.get(selected_column_1)
            encoded_feature_2 = encoded_dict.get(selected_column_2)
            feature_dict[pill_id].append(encoded_feature_1)
            feature_dict[pill_id].append(encoded_feature_2)

    else:
        raise ValueError(f"Wrong dataset_name: {dataset_name}")

    return feature_dict


def process_vectors(sheet, words: list, dataset_name: str, timestamp: str, feature_type: str) -> None:
    """
    Process feature vectors, sort them, and save them to a JSON file.

    Args:
        sheet (Workbook): Excel sheet object.
        words (list): List of words for encoding.
        dataset_name (str): Name of the dataset.
        timestamp (str): Timestamp for the file name.
        feature_type (str): Type of feature vectors.

    Returns:
        None
    """
    
    dictionary = create_feature_vectors(sheet, words, feature_type, dataset_name)
    path = dataset_images_path_selector(dataset_name).get("dynamic_margin").get(f"{feature_type}")
    json_save_filename = os.path.join(path, f"{timestamp}_{feature_type}.json")
    with open(json_save_filename, "w") as json_file:
        json.dump(dictionary, json_file, cls=NumpyEncoder)
    logging.info(f"Saved feature vectors to {json_save_filename}")


def main() -> None:
    """
    
    Returns:
         None
    """
    timestamp = create_timestamp()

    cfg = (
        load_config_json(
            json_schema_filename=json_config_selector("stream_images").get("schema"),
            json_filename=json_config_selector("stream_images").get("config")
        )
    )

    dataset_type = cfg.get('dataset_type')
    imprint_words = ['EMBOSSED', 'PRINTED', 'NOTHING']
    score_words = [1, 2, 4]

    pill_desc_path = dataset_images_path_selector(dataset_type).get("dynamic_margin").get("pill_desc_xlsx")
    pill_desc_file = os.path.join(pill_desc_path, f"pill_desc_{dataset_type}.xlsx")

    workbook = openpyxl.load_workbook(pill_desc_file)
    sheet = workbook['Sheet1']

    process_vectors(sheet, imprint_words, dataset_type, timestamp, "imprint_vectors")
    process_vectors(sheet, score_words, dataset_type, timestamp, "score_vectors")


if __name__ == '__main__':
    main()
